import pandas as pd
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.chart import BarChart, Reference

# ईमेल भेजने के लिए जरूरी मॉड्यूल्स
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

# ==========================================
# 1. डेटा तैयार करना और एक्सेल बनाना (आपका पुराना कोड)
# ==========================================
data = {
    'तारीख': ['13-Jul-2026', '13-Jul-2026', '13-Jul-2026', '13-Jul-2026', '13-Jul-2026'],
    'आइटम': ['Laptop', 'Mouse', 'Keyboard', 'Monitor', 'Mouse'],
    'मात्रा': [2, 5, 3, 1, 2],
    'कीमत': [45000, 600, 1200, 15000, 600],
    'शहर': ['Delhi', 'Mumbai', 'Delhi', 'Bangalore', 'Delhi']
}

df = pd.DataFrame(data)
df['कुल कमाई'] = df['मात्रा'] * df['कीमत']
summary = df.groupby('आइटम')[['मात्रा', 'कुल कमाई']].sum().reset_index()

excel_filename = 'final_sales_report.xlsx'
with pd.ExcelWriter(excel_filename, engine='openpyxl') as writer:
    df.to_excel(writer, sheet_name='Sales Data', index=False)
    summary.to_excel(writer, sheet_name='Summary', index=False)
    
    workbook = writer.book
    
    # स्टाइलिंग
    header_fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    data_font = Font(name="Calibri", size=11)
    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'),
                         top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    
    for sheet_name in ['Sales Data', 'Summary']:
        ws = workbook[sheet_name]
        ws.views.sheetView[0].showGridLines = True
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center")
            
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
            for cell in row:
                cell.font = data_font
                cell.border = thin_border
                if isinstance(cell.value, (int, float)):
                    cell.alignment = Alignment(horizontal="right")
                    if cell.column in [4, 6] or (sheet_name == 'Summary' and cell.column == 3):
                        cell.number_format = '"₹"#,##0'
                else:
                    cell.alignment = Alignment(horizontal="center")
                    
        for col in ws.columns:
            max_len = max(len(str(cell.value or '')) for cell in col)
            col_letter = get_column_letter(col[0].column)
            ws.column_dimensions[col_letter].width = max(max_len + 3, 12)

    # चार्ट जोड़ना
    ws_summary = workbook['Summary']
    chart = BarChart()
    chart.type = "col"
    chart.style = 10
    chart.title = "आइटम के हिसाब से कुल कमाई"
    chart.y_axis.title = "कमाई (₹)"
    chart.x_axis.title = "आइटम"
    
    data_ref = Reference(ws_summary, min_col=3, min_row=1, max_row=ws_summary.max_row)
    cats_ref = Reference(ws_summary, min_col=1, min_row=2, max_row=ws_summary.max_row)
    chart.add_data(data_ref, titles_from_data=True)
    chart.set_categories(cats_ref)
    chart.legend = None
    ws_summary.add_chart(chart, "E2")

print(f"🚀 एक्सेल फाइल तैयार है!")

# ==========================================
# 2. 📨 ईमेल ऑटोमेशन स्टेप
# ==========================================

# ⚠️ यहाँ अपनी डिटेल्स डालें
SENDER_EMAIL = "your_email@gmail.com"      # आपका ईमेल
SENDER_PASSWORD = "your_app_password"     # आपका Gmail App Password (नॉर्मल पासवर्ड नहीं)
RECEIVER_EMAIL = "receiver_email@gmail.com" # जिसे भेजना है उसका ईमेल

print("📨 ईमेल भेजने की प्रक्रिया शुरू हो रही है...")

# ईमेल मैसेज सेटअप करना
msg = MIMEMultipart()
msg['From'] = SENDER_EMAIL
msg['To'] = RECEIVER_EMAIL
msg['Subject'] = "📊 डेली सेल्स रिपोर्ट - 13-Jul-2026"

# ईमेल की बॉडी (Text)
body = """
नमस्ते,

आज की सेल्स रिपोर्ट तैयार हो गई है और इस ईमेल के साथ अटैच कर दी गई है।
इस रिपोर्ट में सेल्स समरी और एक विजुअल बार चार्ट भी शामिल है।

धन्यवाद,
आपका ऑटोमेटेड पायथन बॉट 🤖
"""
msg.attach(MIMEText(body, 'plain'))

# एक्सेल फाइल को अटैच करना
with open(excel_filename, "rb") as attachment:
    part = MIMEBase("application", "octet-stream")
    part.set_payload(attachment.read())
    encoders.encode_base64(part)
    part.add_header(
        "Content-Disposition",
        f"attachment; filename= {excel_filename}",
    )
    msg.attach(part)

# SMTP सर्वर के जरिए ईमेल भेजना
try:
    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls() # सिक्योरिटी इनेबल करें
    server.login("mohammadraja26@gmail.com", "jnnxnfiiyiktcezi")
    text = msg.as_string()
    server.sendmail(SENDER_EMAIL, "reyazulmohammad594@gmail.com", text)
    server.quit()
    print("🎉 बधाई हो! रिपोर्ट ईमेल के जरिए सफलतापूर्वक भेज दी गई है।")
except Exception as e:
    print(f"❌ ईमेल भेजने में गड़बड़ हुई: {e}")